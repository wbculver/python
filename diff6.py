import difflib
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm
import concurrent.futures

def preprocess_ip_route_output(file_path):
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()

            # Remove lines that start with a timestamp-like format
            cleaned_lines = [line for line in lines if not re.match(r'\s*\b\d+[wdhms]\b', line)]

            return cleaned_lines
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return []

def compare_and_create_excel_parallel(file1_path, file2_path, excel_output_path):
    try:
        content1 = preprocess_ip_route_output(file1_path)
        content2 = preprocess_ip_route_output(file2_path)

        differ = difflib.Differ()
        diff = list(differ.compare(content1, content2))

        # Create a new Excel workbook and get the active sheet
        wb = Workbook()
        ws = wb.active

        # Define color for differences
        diff_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        def process_line(line):
            if line.startswith('+') or line.startswith('-'):
                ws.append([line[1:]])
                ws.cell(row=ws.max_row, column=1).fill = diff_fill

        # Use concurrent.futures for parallel processing
        with concurrent.futures.ThreadPoolExecutor() as executor:
            list(tqdm(executor.map(process_line, diff), total=len(diff), desc="Comparing and Creating Excel"))

        # Save the Excel file
        wb.save(excel_output_path)
    except FileNotFoundError:
        print("One or both files not found.")

# Provide the paths of the two files and the Excel output file
file1_path = 'before.txt'
file2_path = 'after.txt'
excel_output_path = 'diff6.xlsx'

compare_and_create_excel_parallel(file1_path, file2_path, excel_output_path)
