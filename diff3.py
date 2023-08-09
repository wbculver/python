import difflib
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def compare_and_create_excel(file1_path, file2_path, excel_output_path):
    try:
        with open(file1_path, 'r') as file1, open(file2_path, 'r') as file2:
            content1 = file1.readlines()
            content2 = file2.readlines()

            differ = difflib.Differ()
            diff = list(differ.compare(content1, content2))

            # Create a new Excel workbook and get the active sheet
            wb = Workbook()
            ws = wb.active

            # Define color for differences
            diff_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Iterate through the diff and write to Excel with color
            for line in diff:
                if line.startswith('+') or line.startswith('-'):
                    ws.append([line[1:]])
                    ws.cell(row=ws.max_row, column=1).fill = diff_fill
                elif line.startswith(' '):
                    ws.append([line[1:]])

            # Save the Excel file
            wb.save(excel_output_path)
    except FileNotFoundError:
        print("One or both files not found.")

# Provide the paths of the two files and the Excel output file
file1_path = 'before.txt'
file2_path = 'after.txt'
excel_output_path = 'diff.xlsx'

compare_and_create_excel(file1_path, file2_path, excel_output_path)
