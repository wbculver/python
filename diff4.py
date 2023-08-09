import difflib
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def preprocess_ip_route_output(file_path):
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()

            # Remove the timestamp using regex
            cleaned_lines = [re.sub(r'\b\d+[wdhm]\d+[wdhm]\b', '', line) for line in lines]

            return cleaned_lines
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return []

def compare_and_create_excel(file1_path, file2_path, excel_output_path):
    try:
        content1 = preprocess_ip_route_output(file1_path)
        content2 = preprocess_ip_route_output(file2_path)

        differ = difflib.Differ()
        diff = list(differ.compare(content1, content2))

        # Create a new Excel workbook and get the active sheet
        wb = Workbook()
        ws = wb.active

        # Define color for differences
        diff_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Iterate through the diff and write only changed lines to Excel with color
        for line in diff:
            if line.startswith('+') or line.startswith('-'):
                ws.append([line[1:]])
                ws.cell(row=ws.max_row, column=1).fill = diff_fill

        # Save the Excel file
        wb.save(excel_output_path)
    except FileNotFoundError:
        print("One or both files not found.")

# Provide the paths of the two files and the Excel output file
file1_path = 'path/to/first/file.txt'
file2_path = 'path/to/second/file.txt'
excel_output_path = 'path/to/output/diff.xlsx'

compare_and_create_excel(file1_path, file2_path, excel_output_path)
