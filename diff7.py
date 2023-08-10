import difflib
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm

def preprocess_ip_route_output(file_path):
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()

            # Remove lines that start with a timestamp-like format
            cleaned_lines = [line for line in lines if not re.match(r'\s*\b\d+[wdhms]\b', line)]

            return cleaned_lines
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return []

def compare_and_create_excel(file1_path, file2_path, excel_output_path):
    try:
        content1 = preprocess_ip_route_output(file1_path)
        content2 = preprocess_ip_route_output(file2_path)

        differ = difflib.Differ()
        diff = list(differ.compare(content1, content2))

        # Create a new Excel workbook and get the active sheet
        wb = Workbook()
        ws = wb.active

        # Define colors for differences
        added_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        removed_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

        for line in tqdm(diff, desc="Comparing and Creating Excel"):
            if line.startswith('+'):
                ws.append([line[1:]])
                ws.cell(row=ws.max_row, column=1).fill = added_fill
            elif line.startswith('-'):
                ws.append([line[1:]])
                ws.cell(row=ws.max_row, column=1).fill = removed_fill

        # Save the Excel file
        wb.save(excel_output_path)
    except FileNotFoundError:
        print("One or both files not found.")

# Provide the paths of the two files and the Excel output file
file1_path = 'path/to/first/file.txt'
file2_path = 'path/to/second/file.txt'
excel_output_path = 'path/to/output/diff.xlsx'

compare_and_create_excel(file1_path, file2_path, excel_output_path)
